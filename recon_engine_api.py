import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

_SALES_HINTS = {
    "invoice_no":      ["inv no","invoice number","invoice no","bill no","doc no","doc.no"],
    "invoice_date":    ["inv date","invoice date","bill date","date"],
    "invoice_value":   ["grand total","invoice value","total invoice value","total value","bill value","inv value"],
    "taxable_value":   ["acc value","taxable value","assessable value","taxable amount","taxable amt"],
    "gstin":           ["gst no","gstin/uin","gstin of recipient","recipient gstin","party gstin","gstin"],
    "receiver_name":   ["party name","receiver name","customer name","buyer name"],
    "place_of_supply": ["place of supply","pos","destination state"],
    "igst":            ["igst amount","integrated tax","igst"],
    "cgst":            ["cgst amount","central tax","cgst"],
    "sgst":            ["sgst amount","state / ut tax","sgst","ut tax"],
    "rate":            ["rate","gst rate","tax rate"],
    "doc_type":        ["doctype","doc type","document type","doc.type"],
}

_EWB_HINTS = {
    "doc_no":      ["doc.no","doc no","document no","invoice no","bill no"],
    "doc_date":    ["doc.date","doc date","document date"],
    "ewb_no":      ["ewb no","e-way bill no","eway bill no","ewb number"],
    "ewb_date":    ["ewb date","e-way bill date","eway bill date","generated date"],
    "supply_type": ["supply type"],
    "doc_type":    ["doc.type","doc type","document type"],
    "from_gstin":  ["from gstin"],
    "to_gstin":    ["to gstin"],
    "from_info":   ["from gstin info"],
    "to_info":     ["to gstin info"],
    "total_value": ["total invoice value","total value","assessable value"],
    "status":      ["status"],
    "hsn":         ["main hsn code","hsn code","hsn"],
    "irn":         ["irn"],
}

def _col(df, hints):
    for h in hints:
        for c in df.columns:
            if h.lower() == str(c).lower().strip(): return c
    for h in hints:
        for c in df.columns:
            if h.lower() in str(c).lower(): return c
    return None

def _norm(val):
    if val is None: return ""
    try:
        if pd.isna(val): return ""
    except Exception: pass
    s = str(val).strip()
    if s.endswith(".0"): s = s[:-2]
    return s.upper()

def _flt(val):
    try:
        if pd.isna(val): return 0.0
    except Exception: pass
    try:
        return float(str(val).replace(",","").replace("₹","").strip())
    except Exception:
        return 0.0

def _is_cdn(key, val, doc_type=""):
    if val < 0: return True
    dt = str(doc_type).upper().strip()
    if any(x in dt for x in ("CREDIT","DEBIT","CDN","CN NOTE","DN NOTE")): return True
    for p in ("CN","CR","CDN","DN","DR","CDNR"):
        if key.startswith(p): return True
    return False

def probe_file_for_claude(file_bytes):
    xl = pd.ExcelFile(BytesIO(file_bytes))
    result = {"sheet_names": xl.sheet_names, "sheets": {}}
    for sheet in xl.sheet_names:
        try:
            df_full = xl.parse(sheet)
            df_head = xl.parse(sheet, nrows=3)
            result["sheets"][sheet] = {
                "columns":    [str(c) for c in df_full.columns],
                "row_count":  len(df_full),
                "sample_rows": [{str(k): str(v) for k,v in row.items()}
                                for row in df_head.fillna("").to_dict("records")],
            }
        except Exception as ex:
            result["sheets"][sheet] = {"columns":[],"row_count":0,"sample_rows":[],"error":str(ex)}
    return result

def run_reconciliation(file_bytes, config, ewb_file_bytes=None, einv_file_bytes=None):
    sheet_s   = config.get("sheet_sales",      "sales")
    sheet_e   = config.get("sheet_ewb",        "e way bill")
    cli_name  = config.get("client_name",      "")
    gstin     = config.get("gstin",            "")
    period    = config.get("tax_period",       "")
    cdn_row   = config.get("cdn_start_row",    None)
    hdr_row   = config.get("sales_header_row", 0)
    einv_on   = config.get("einv_applicable",  False)

    def find_sheet(xl, name):
        for s in xl.sheet_names:
            if s.strip().lower() == name.strip().lower(): return s
        for s in xl.sheet_names:
            if name.strip().lower() in s.strip().lower(): return s
        return xl.sheet_names[0]

    xl_s  = pd.ExcelFile(BytesIO(file_bytes))
    sales = xl_s.parse(find_sheet(xl_s, sheet_s), header=hdr_row)
    sales.columns = [str(c).strip() for c in sales.columns]

    if ewb_file_bytes:
        xl_e = pd.ExcelFile(BytesIO(ewb_file_bytes))
        ewb  = xl_e.parse(find_sheet(xl_e, sheet_e))
    else:
        ewb  = xl_s.parse(find_sheet(xl_s, sheet_e))
    ewb.columns = [str(c).strip() for c in ewb.columns]

    einv = None
    if einv_on and einv_file_bytes:
        xl_i = pd.ExcelFile(BytesIO(einv_file_bytes))
        einv = xl_i.parse(xl_i.sheet_names[0])
        einv.columns = [str(c).strip() for c in einv.columns]

    s = {k: _col(sales, h) for k, h in _SALES_HINTS.items()}
    e = {k: _col(ewb,   h) for k, h in _EWB_HINTS.items()}

    if not cli_name and e["from_info"]:
        raw = ewb[e["from_info"]].dropna()
        cli_name = str(raw.iloc[0]).split("  ")[0].strip() if len(raw) else ""
    if not gstin and e["from_gstin"]:
        raw = ewb[e["from_gstin"]].dropna()
        gstin = str(raw.iloc[0]).strip() if len(raw) else ""

    if s["invoice_no"]:
        sales = sales[sales[s["invoice_no"]].notna()].copy()
        sales = sales[sales[s["invoice_no"]].astype(str).str.strip() != ""].copy()
        sales = sales[~sales[s["invoice_no"]].astype(str).str.lower().str.contains(
            "inv no|invoice no|total|subtotal", na=False)].copy()

    sales["_key"] = sales[s["invoice_no"]].apply(_norm) if s["invoice_no"] else ""
    ewb["_key"]   = ewb[e["doc_no"]].apply(_norm)       if e["doc_no"]     else ""
    sales["_val"] = sales[s["invoice_value"]].apply(_flt) if s["invoice_value"] else 0.0

    if e["doc_type"]:
        tax_mask    = ewb[e["doc_type"]].astype(str).str.lower().str.contains(
                          "tax invoice|invoice", na=False)
        ewb_tax     = ewb[tax_mask].copy()
        ewb_challan = ewb[~tax_mask].copy()
    else:
        ewb_tax     = ewb.copy()
        ewb_challan = pd.DataFrame()

    if e["status"]:
        ewb_tax = ewb_tax[~ewb_tax[e["status"]].astype(str).str.lower().str.contains(
            "cancel", na=False)]
    if e["supply_type"]:
        ewb_tax = ewb_tax[ewb_tax[e["supply_type"]].astype(str).str.lower().str.contains(
            "outward", na=False)]

    doc_type_col = s.get("doc_type")
    sales["_cdn"] = sales.apply(
        lambda r: _is_cdn(r["_key"], r["_val"],
                          r.get(doc_type_col, "") if doc_type_col else ""), axis=1)
    if cdn_row is not None:
        sales.loc[sales.index >= cdn_row - 2, "_cdn"] = True
    cdn_df  = sales[sales["_cdn"]].copy()
    sales_m = sales[~sales["_cdn"]].copy()

    ewb_lkp = {r["_key"]: r for _, r in ewb_tax.iterrows()}
    matched, no_ewb, mismatch = [], [], []

    for _, sr in sales_m.iterrows():
        key = sr["_key"]
        if not key: continue
        if key in ewb_lkp:
            er  = ewb_lkp[key]
            sv  = sr["_val"]
            ev  = _flt(er.get(e["total_value"], 0)) if e["total_value"] else 0.0
            tol = max(1.0, abs(sv) * 0.001)
            if abs(sv - ev) <= tol or sv == 0:
                matched.append({"s": sr, "e": er})
            else:
                mismatch.append({"s": sr, "e": er, "sv": sv, "ev": ev})
        else:
            no_ewb.append(sr)

    used_keys = {r["s"]["_key"] for r in matched} | {r["s"]["_key"] for r in mismatch}
    orphan    = ewb_tax[~ewb_tax["_key"].isin(used_keys)].copy()

    einv_missing = []
    if einv is not None:
        irn_col    = next((c for c in einv.columns if "irn" in c.lower()), None)
        inv_col    = next((c for c in einv.columns if any(
                           x in c.lower() for x in ["doc no","invoice no","inv no"])), None)
        cancel_col = next((c for c in einv.columns if "cancel" in c.lower()), None)
        if irn_col and inv_col:
            einv_active = einv.copy()
            if cancel_col:
                einv_active = einv_active[
                    ~einv_active[cancel_col].astype(str).str.lower().str.contains(
                        "cancel|y", na=False)]
            einv_keys = set(einv_active[inv_col].apply(_norm))
            for sr in no_ewb + [r["s"] for r in matched] + [r["s"] for r in mismatch]:
                k = sr["_key"]
                if k and k not in einv_keys:
                    einv_missing.append(sr)

    total_s  = len(sales_m)
    total_e  = len(ewb_tax)
    ok       = len(matched)
    vm       = len(mismatch)
    no_e     = len(no_ewb)
    orph     = len(orphan)
    cdn_cnt  = len(cdn_df)
    challan  = len(ewb_challan)
    einv_gap = len(einv_missing)
    issues   = vm + no_e + orph + (einv_gap if einv_on else 0)
    rate     = round(ok / total_s * 100, 1) if total_s else 0.0

    stats = {
        "client_name":       cli_name,
        "gstin":             gstin,
        "tax_period":        period,
        "total_sales":       total_s,
        "total_ewb":         total_e,
        "delivery_challans": challan,
        "matched_ok":        ok,
        "val_mismatch":      vm,
        "sales_no_ewb":      no_e,
        "ewb_no_sales":      orph,
        "cdnr_count":        cdn_cnt,
        "einv_missing_irn":  einv_gap,
        "match_rate_pct":    rate,
        "issues_count":      issues,
    }

    wb  = Workbook()
    wb.remove(wb.active)
    H_FILL = PatternFill("solid", fgColor="1F4E79")
    H_FONT = Font(color="FFFFFF", bold=True, size=10)
    OK_F   = PatternFill("solid", fgColor="E2EFDA")
    WN_F   = PatternFill("solid", fgColor="FFF2CC")
    ER_F   = PatternFill("solid", fgColor="FCE4D6")
    BRD    = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))

    def hdr(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.fill=H_FILL; c.font=H_FONT
            c.alignment=Alignment(horizontal="center", wrap_text=True)
            c.border=BRD

    def aw(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+2, 45)

    def get(row, key, col_map):
        c = col_map.get(key)
        return row.get(c, "") if c else ""

    def ewb_dt(er):
        if not e["ewb_date"]: return ""
        try:
            d = pd.to_datetime(er.get(e["ewb_date"],""), errors="coerce")
            return str(d.date()) if pd.notna(d) else ""
        except Exception: return ""

    ws1 = wb.create_sheet("Summary")
    filing_verdict = ("Hold — do not file" if rate < 85 else
                      "Review required"    if rate < 95 else "Ready to file")
    for row in [
        ["GSTR-1 Reconciliation Report",""],
        ["Client", cli_name],["GSTIN", gstin],["Tax Period", period],
        ["Run Date", datetime.now().strftime("%d-%b-%Y %H:%M")],["",""],
        ["Total Sales (Tax Invoices)", total_s],
        ["Total EWBs (Tax Invoices)",  total_e],
        ["Delivery Challans (Job Work)", challan],
        ["Matched", ok],["Value Mismatch", vm],
        ["Sales No EWB", no_e],["EWB No Match", orph],
        ["Credit Debit Notes", cdn_cnt],
        ["E-Invoice IRN Missing", einv_gap],["",""],
        ["Match Rate %", f"{rate}%"],
        ["Total Issues", issues],
        ["Filing Verdict", filing_verdict],
    ]:
        ws1.append(row)
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 45

    ws2 = wb.create_sheet("Query Sheet")
    hdr(ws2, ["#","Issue Type","Invoice No","Invoice Date","Invoice Value",
              "Party Name","GSTIN","Place of Supply",
              "EWB No","EWB Date","EWB Value","Remarks"])
    n = 0

    for sr in no_ewb:
        n += 1
        ws2.append([n,"No EWB Found",
                    get(sr,"invoice_no",s), get(sr,"invoice_date",s),
                    get(sr,"invoice_value",s), get(sr,"receiver_name",s),
                    get(sr,"gstin",s), get(sr,"place_of_supply",s),
                    "","","","Sales invoice has no matching EWB Tax Invoice"])
        for c in ws2[ws2.max_row]: c.fill=WN_F; c.border=BRD

    for item in mismatch:
        n += 1
        sr, er = item["s"], item["e"]
        ws2.append([n,"Value Mismatch",
                    get(sr,"invoice_no",s), get(sr,"invoice_date",s), item["sv"],
                    get(sr,"receiver_name",s), get(sr,"gstin",s),
                    get(sr,"place_of_supply",s),
                    er.get(e["ewb_no"],"") if e["ewb_no"] else "",
                    ewb_dt(er), item["ev"],
                    "Sales vs EWB value difference"])
        for c in ws2[ws2.max_row]: c.fill=ER_F; c.border=BRD

    for _, er in orphan.iterrows():
        n += 1
        ws2.append([n,"EWB No Sales Match",
                    er.get(e["doc_no"],"")     if e["doc_no"]     else "",
                    er.get(e["doc_date"],"")   if e["doc_date"]   else "",
                    er.get(e["total_value"],"") if e["total_value"] else "",
                    er.get(e["to_info"],"")    if e["to_info"]    else "",
                    er.get(e["to_gstin"],"")   if e["to_gstin"]   else "","",
                    er.get(e["ewb_no"],"")     if e["ewb_no"]     else "",
                    ewb_dt(er),
                    er.get(e["total_value"],"") if e["total_value"] else "",
                    "EWB Tax Invoice has no matching sales entry"])
        for c in ws2[ws2.max_row]: c.fill=ER_F; c.border=BRD

    if einv_on:
        for sr in einv_missing:
            n += 1
            ws2.append([n,"E-Invoice IRN Missing",
                        get(sr,"invoice_no",s), get(sr,"invoice_date",s),
                        get(sr,"invoice_value",s), get(sr,"receiver_name",s),
                        get(sr,"gstin",s), get(sr,"place_of_supply",s),
                        "","","","Invoice has no IRN in e-invoice register"])
            for c in ws2[ws2.max_row]: c.fill=ER_F; c.border=BRD
    aw(ws2)

    ws3 = wb.create_sheet("Matched")
    hdr(ws3, ["Invoice No","Invoice Date","Invoice Value",
              "Party Name","GSTIN","Place of Supply",
              "EWB No","EWB Date","EWB Value","Match"])
    for item in matched:
        sr, er = item["s"], item["e"]
        ws3.append([get(sr,"invoice_no",s), get(sr,"invoice_date",s),
                    get(sr,"invoice_value",s), get(sr,"receiver_name",s),
                    get(sr,"gstin",s), get(sr,"place_of_supply",s),
                    er.get(e["ewb_no"],"") if e["ewb_no"] else "",
                    ewb_dt(er),
                    er.get(e["total_value"],"") if e["total_value"] else "","✔"])
        for c in ws3[ws3.max_row]: c.fill=OK_F; c.border=BRD
    aw(ws3)

    if len(ewb_challan):
        ws4 = wb.create_sheet("Delivery Challans")
        hdr(ws4, ["Doc No","Doc Date","Doc Type","From GSTIN",
                  "To GSTIN","To Party","Total Value","EWB No","Status"])
        for _, cr in ewb_challan.iterrows():
            ws4.append([
                cr.get(e["doc_no"],"")      if e["doc_no"]      else "",
                cr.get(e["doc_date"],"")    if e["doc_date"]    else "",
                cr.get(e["doc_type"],"")    if e["doc_type"]    else "",
                cr.get(e["from_gstin"],"")  if e["from_gstin"]  else "",
                cr.get(e["to_gstin"],"")    if e["to_gstin"]    else "",
                cr.get(e["to_info"],"")     if e["to_info"]     else "",
                cr.get(e["total_value"],"") if e["total_value"] else "",
                cr.get(e["ewb_no"],"")      if e["ewb_no"]      else "",
                cr.get(e["status"],"")      if e["status"]      else "",
            ])
            for c in ws4[ws4.max_row]: c.fill=WN_F; c.border=BRD
        aw(ws4)

    if len(cdn_df):
        ws5 = wb.create_sheet("CDNR")
        hdr(ws5, ["Invoice No","Invoice Date","Invoice Value",
                  "Party Name","GSTIN","Place of Supply"])
        for _, cr in cdn_df.iterrows():
            ws5.append([get(cr,"invoice_no",s), get(cr,"invoice_date",s),
                        get(cr,"invoice_value",s), get(cr,"receiver_name",s),
                        get(cr,"gstin",s), get(cr,"place_of_supply",s)])
            for c in ws5[ws5.max_row]: c.fill=WN_F; c.border=BRD
        aw(ws5)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), stats
